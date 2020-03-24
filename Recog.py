import cv2
from align_custom import AlignCustom
from face_feature import FaceFeature
from mtcnn_detect import MTCNNDetect
from tf_graph import FaceRecGraph
import openpyxl as xl
import datetime
import time
import argparse
import sys
import json
import numpy as np
import os 

def main(args):
    mode = args.mode
    if(mode == "camera"):
        camera_recog()
    else:
        raise ValueError("Unimplemented mode")
def camera_recog():
    # Load present date and time
    now= datetime.datetime.now()  
    today=now.day
    month=now.month

    currentDate = time.strftime("%d_%m_%y")

    path1 = r'Class.xlsx'
    path2 = r'Attendance.xlsx'
    names = json.load(open("names.txt"))
 

    wb1 = xl.load_workbook(filename=path1)
    ws1 = wb1.worksheets[0]

    wb2 = xl.load_workbook(filename=path2)
    ws2 = wb2.worksheets[0]
    # ws2 = wb2.create_sheet('Att'+str(today))

    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

   
    print("[INFO] camera sensor warming up...")
    vs = cv2.VideoCapture(0); #get input from webcam
    while True:
        _,frame = vs.read();
        #u can certainly add a roi here but for the sake of a demo i'll just leave it as simple as this
        rects, landmarks = face_detect.detect_face(frame,80);#min face size is set to 80x80
        aligns = []
        positions = []
        

    
        for (i, rect) in enumerate(rects):
            aligned_face, face_pos = aligner.align(160,frame,landmarks[i])
            if len(aligned_face) == 160 and len(aligned_face[0]) == 160:
                aligns.append(aligned_face)
                positions.append(face_pos)
            else: 
                print("Align face failed") #log        
        if(len(aligns) > 0):
            features_arr = extract_feature.get_features(aligns)
            recog_data = findPeople(features_arr,positions);
            for (i,rect) in enumerate(rects):
                cv2.rectangle(frame,(rect[0],rect[1]),(rect[0] + rect[2],rect[1]+rect[3]),(0,255,0),2) #draw bounding box for the face
                cv2.putText(frame,recog_data[i][0]+" - "+str(recog_data[i][1])+"%",(rect[0],rect[1]),cv2.FONT_HERSHEY_SIMPLEX,1,(255,255,255),1,cv2.LINE_AA)
                for rollno, name in names.items():   
                   if name == recog_data[i][0]:
                       ws2.cell(row=int(1), column=int(today)).value = currentDate
                       ws2.cell(row=int(rollno), column=int(today)).value = "Present"
                   else:
                       pass 
                  
        cv2.imshow("Capturing Face",frame)
        key = cv2.waitKey(1) & 0xFF
        if key == 27 or key == ord("q"):
            break
    vs.release() # camera release 
    cv2.destroyAllWindows()  
        # Save Woorksheet
    wb2.save(path2)
    # os.system('python Dupli.py')
    # os.system('python Readonly.py')

'''
facerec_128D.txt Data Structure:
{
"Person ID": {
    "Center": [[128D vector]],
    "Left": [[128D vector]],
    "Right": [[128D Vector]]
    }
}
This function basically does a simple linear search for 
^the 128D vector with the min distance to the 128D vector of the face on screen
'''
def findPeople(features_arr, positions, thres = 0.6, percent_thres = 70):
    '''
    :param features_arr: a list of 128d Features of all faces on screen
    :param positions: a list of face position types of all faces on screen
    :param thres: distance threshold
    :return: person name and percentage
    '''
    f = open('./facerec_128D.txt','r')
    data_set = json.loads(f.read());
    returnRes = [];
    for (i,features_128D) in enumerate(features_arr):
        result = "Unknown";
        smallest = sys.maxsize
        for person in data_set.keys():
            person_data = data_set[person][positions[i]];
            for data in person_data:
                distance = np.sqrt(np.sum(np.square(data-features_128D)))
                if(distance < smallest):
                    smallest = distance;
                    result = person;
        percentage =  min(100, 100 * thres / smallest)
        if percentage <= percent_thres :
            result = "Unknown"
        returnRes.append((result,percentage))
    return returnRes

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", type=str, help="Run camera recognition", default="camera")
    args = parser.parse_args(sys.argv[1:]);
    FRGraph = FaceRecGraph();
    aligner = AlignCustom();
    extract_feature = FaceFeature(FRGraph)
    face_detect = MTCNNDetect(FRGraph, scale_factor=2); #scale_factor, rescales image for faster detection
    main(args);